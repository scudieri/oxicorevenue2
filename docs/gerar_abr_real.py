import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import date, timedelta

wb = openpyxl.Workbook()

# ── Paleta ───────────────────────────────────────────────────
BLACK    = "0D0D0D"; DARK  = "1A1A1A"; MID   = "252525"; CARD  = "2A2A2A"
ACCENT   = "C8FF00"; CYAN  = "00E5FF"; WHITE = "FFFFFF"; DIM   = "666666"
G_BG     = "1A3320"; R_BG  = "3A1A1A"; O_BG  = "2A2010"; Y_BG  = "2A2200"
BLUE_BG  = "0D1A2A"

def fill(c): return PatternFill("solid", fgColor=c)
def fnt(bold=False, color=WHITE, size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")
def bdr(color="303030"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)
def aln(h="left", wrap=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

def c(ws, r, col, v="", bg=CARD, fg=WHITE, bold=False, fmt=None, sz=10, align="left"):
    cl = ws.cell(r, col, v)
    cl.fill = fill(bg); cl.font = fnt(bold, fg, sz)
    cl.alignment = aln(align, wrap=True); cl.border = bdr()
    if fmt: cl.number_format = fmt
    return cl

def hdr(ws, r, col, text, bg=DARK, fg=ACCENT, sz=9):
    cl = ws.cell(r, col, text)
    cl.fill = fill(bg); cl.font = fnt(True, fg, sz)
    cl.alignment = aln("center"); cl.border = bdr()
    ws.row_dimensions[r].height = 20

def banner(ws, r, text, ncols, bg=BLACK, fg=ACCENT, sz=13):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    cl = ws.cell(r, 1, text)
    cl.fill = fill(bg); cl.font = fnt(True, fg, sz)
    cl.alignment = aln("center"); ws.row_dimensions[r].height = 28

def set_cols(ws, widths):
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def section_row(ws, r, text, ncols, fg=CYAN):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    cl = ws.cell(r, 1, text)
    cl.fill = fill(BLACK); cl.font = fnt(True, fg, 9)
    cl.alignment = aln("center"); cl.border = bdr()
    ws.row_dimensions[r].height = 14

def kpi(ws, r, col, label, value, fg_val=ACCENT, fmt=None):
    cl = ws.cell(r, col, label)
    cl.fill = fill(DARK); cl.font = fnt(False, DIM, 8)
    cl.alignment = aln("center", wrap=True); cl.border = bdr()
    ws.row_dimensions[r].height = 14
    cl2 = ws.cell(r+1, col, value)
    cl2.fill = fill(MID); cl2.font = fnt(True, fg_val, 13)
    cl2.alignment = aln("center"); cl2.border = bdr()
    if fmt: cl2.number_format = fmt
    ws.row_dimensions[r+1].height = 24

# ── Dias úteis ABR/2026 (usando 22 conforme planilha original) ──
# A planilha original diz 22 dias úteis. Inclui 21/04 (Tiradentes).
DIAS_UTEIS = []
d = date(2026, 4, 1)
while d.month == 4:
    if d.weekday() < 5:
        DIAS_UTEIS.append(d)
    d += timedelta(1)
# = 22 dias


# ════════════════════════════════════════════════════════════
# ABA 1 — RESUMO (KPIs do mês)
# ════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "📊 Resumo"
ws1.sheet_properties.tabColor = ACCENT

banner(ws1, 1, "CONTROLE DE RESULTADOS  —  ABRIL 2026", 8)

# Linha KPIs principais
kpis_top = [
    ("META DO MÊS",          235000,  ACCENT, 'R$ #.##0,00'),
    ("OPORTUNIDADES",        502778.93, CYAN,  'R$ #.##0,00'),
    ("IDEAL DIA",            605303.03, DIM,   'R$ #.##0,00'),
    ("GAP",                 -102524.10, "FF4444", 'R$ #.##0,00'),
    ("DIAS ÚTEIS MÊS",       22,        WHITE, '0'),
    ("DIAS ÚTEIS HOJE",      17,        WHITE, '0'),
    ("RITMO MARCAÇÕES",      63,        WHITE, '0'),
    ("RITMO ACONTECIDAS",    50,        WHITE, '0'),
]
for i, (lbl, val, fg, fmt) in enumerate(kpis_top, 1):
    kpi(ws1, 2, i, lbl, val, fg_val=fg, fmt=fmt)

# Linha 4-5 já usada por kpi (label row=2, value row=3)
r = 5
section_row(ws1, r, "PERFORMANCE CLOSERS", 8)
r += 1

# Tabela closers
closer_hdrs = ["CLOSER", "REUNIÕES MARCADAS", "REUNIÕES REALIZADAS", "CONV %", "VENDAS", "TICKET MÉDIO", "RECEITA", "TX SHOW"]
for ci, h in enumerate(closer_hdrs, 1):
    hdr(ws1, r, ci, h)
r += 1

closers_data = [
    ("João",        12, 12, "25%", 3,  11175.00,   33525.00,   "100%"),
    ("Bruno",       30, 20, "20%", 4,  8284.73,    33138.91,   "66,67%"),
    ("Luis Felipe",  0,  0,  "—",  0,  0,          0,           "—"),
    ("TOTAL",       42, 32, "22%", 7,  9523.42,    66663.91,   "83,33%"),
]
for row_d in closers_data:
    is_total = row_d[0] == "TOTAL"
    bg = MID if not is_total else DARK
    fg = ACCENT if is_total else WHITE
    for ci, v in enumerate(row_d, 1):
        fmt = 'R$ #.##0,00' if ci in (6, 7) else None
        c(ws1, r, ci, v, bg=bg, fg=fg, bold=is_total, fmt=fmt, align="center")
    r += 1

r += 1
section_row(ws1, r, "METAS SDR / PACING", 8)
r += 1
sdr_hdrs = ["SDR", "TOTAL SHOW", "TOTAL NO SHOW", "TOTAL MARCADAS", "TOTAL REMARCADAS", "TAXA SHOW %", "META MARCAÇÕES", "GAP MARCAÇÕES"]
for ci, h in enumerate(sdr_hdrs, 1):
    hdr(ws1, r, ci, h)
r += 1
# Totals do SDR (combinado Daniel + Marcelo)
sdr_vals = [("TOTAL", 33, 9, 42, 7, "79%", 63, -21)]
for row_d in sdr_vals:
    for ci, v in enumerate(row_d, 1):
        c(ws1, r, ci, v, bg=MID, fg=ACCENT, bold=True, align="center")
r += 1

r += 1
section_row(ws1, r, "LEADBROKER", 8)
r += 1
lb_hdrs = ["ORÇAMENTO MENSAL", "ORÇAMENTO SEMANAL", "ORÇAMENTO DIÁRIO", "CPV", "NA RUA", "", "", ""]
lb_vals  = [100000, 33333.33, 5000, 3798.63, 0, "", "", ""]
for ci, h in enumerate(lb_hdrs, 1):
    hdr(ws1, r, ci, h)
r += 1
for ci, v in enumerate(lb_vals, 1):
    fmt = 'R$ #.##0,00' if isinstance(v, float) else None
    c(ws1, r, ci, v, bg=MID, fg=ACCENT if v else DIM, bold=True, fmt=fmt, align="center")

set_cols(ws1, [16, 18, 18, 10, 10, 14, 16, 14])


# ════════════════════════════════════════════════════════════
# ABA 2 — PIPELINE (Deals / Leads)
# ════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("🎯 Pipeline")
ws2.sheet_properties.tabColor = "FF9900"

banner(ws2, 1, "PIPELINE DE VENDAS — ABRIL 2026", 10, bg="1A1200", fg="FF9900")

hdrs2 = ["CLIENTE", "VALOR (R$)", "STATUS", "TEMPERATURA", "DATA LEAD", "CANAL", "PRODUTO", "CLOSER", "SDR", "DATA ASSINATURA"]
hints2 = ["nome da empresa", "valor proposto", "Fechado/Proposta/Negativou/Stand by", "Quente/Morno/Frio", "mês de entrada", "LeadBroker/Evento/etc.", "ONE TIME/Ass. Booking", "Bruno/João/Luis Felipe", "Daniel/Marcelo/João/Bruno", "dd/mm/aaaa"]
for ci, h in enumerate(hdrs2, 1):
    hdr(ws2, 2, ci, h, bg="1A1200", fg="FF9900")
for ci, h in enumerate(hints2, 1):
    c(ws2, 3, ci, h, bg="0D0A00", fg=DIM, sz=8, align="center")

# Status color map
STATUS_BG = {
    "Fechado":  G_BG,
    "Proposta": O_BG,
    "Negativou": R_BG,
    "Stand by": BLUE_BG,
}
TEMP_FG = {"Quente": "FF4444", "Morno": "FF9900", "Frio": CYAN}

deals = [
    # (cliente, valor, status, temperatura, data_lead, canal, produto, closer, sdr, data_assinatura)
    ("Foton",                    12364.00,  "Fechado",  "Quente", "Março",  "LeadBroker",        "ONE TIME",    "Bruno",       "Daniel",       "06/04/2026"),
    ("Vie Douce",                3301.01,   "Fechado",  "Quente", "abril",  "LeadBroker",        "Ass. Booking","João",        "Daniel",       ""),
    ("Vie Douce",                5479.92,   "Fechado",  "Quente", "abril",  "LeadBroker",        "ONE TIME",    "João",        "Daniel",       ""),
    ("Atlas",                    14050.00,  "Fechado",  "Quente", "abril",  "LeadBroker",        "ONE TIME",    "Bruno",       "Daniel",       "10/04/2026"),
    ("Home Interiores",          17984.00,  "Fechado",  "Quente", "abril",  "LeadBroker",        "ONE TIME",    "João",        "Daniel",       "15/04/2026"),
    ("Ergon Vision",             2799.00,   "Fechado",  "Quente", "abril",  "LeadBroker",        "ONE TIME",    "João",        "João",         ""),
    ("Ergon Vision",             3961.21,   "Fechado",  "Quente", "abril",  "LeadBroker",        "Ass. Booking","João",        "João",         ""),
    ("Tornearia do Veinho",      3159.82,   "Fechado",  "Quente", "abril",  "LeadBroker",        "Ass. Booking","Bruno",       "Bruno",        "16/04/2026"),
    ("ROTA MAQ",                 3565.09,   "Fechado",  "Quente", "Março",  "LeadBroker",        "Ass. Booking","Bruno",       "João",         ""),
    ("Arkan",                    20000.00,  "Proposta", "Quente", "abril",  "LeadBroker",        "ONE TIME",    "João",        "",             ""),
    ("MELK",                     None,      "Proposta", "Quente", "abril",  "LeadBroker",        "Ass. Booking","Bruno",       "João",         ""),
    ("Essence Tercerização",     None,      "Proposta", "Quente", "abril",  "LeadBroker",        "Ass. Booking","Bruno",       "João",         ""),
    ("La ensenada",              45750.00,  "Proposta", "Morno",  "Março",  "Evento",            "ONE TIME",    "Luis Felipe", "Luis Felipe",  ""),
    ("Indusbello",               64633.80,  "Proposta", "Morno",  "Março",  "Evento",            "ONE TIME",    "Luis Felipe", "Luis Felipe",  ""),
    ("Hyleflex",                 18000.00,  "Proposta", "Morno",  "abril",  "Recuperação Broker","ONE TIME",    "João",        "",             ""),
    ("Joy",                      25000.00,  "Proposta", "Morno",  "abril",  "Recuperação Broker","Ass. Booking","João",        "",             ""),
    ("Arrue",                    14050.00,  "Proposta", "Morno",  "abril",  "LeadBroker",        "ONE TIME",    "Bruno",       "",             ""),
    ("Clinica Amare",            3300.00,   "Proposta", "Morno",  "abril",  "LeadBroker",        "Ass. Booking","Bruno",       "João",         ""),
    ("Confido",                  None,      "Proposta", "Morno",  "abril",  "LeadBroker",        "Ass. Booking","Bruno",       "João",         ""),
    ("Erione",                   None,      "Proposta", "Frio",   "abril",  "LeadBroker",        "ONE TIME",    "Bruno",       "",             ""),
    ("Silvio Brandão",           18000.00,  "Proposta", "Frio",   "abril",  "LeadBroker",        "ONE TIME",    "João",        "",             ""),
    ("Advan",                    21000.00,  "Proposta", "Frio",   "abril",  "LeadBroker",        "Ass. Booking","João",        "",             ""),
    ("Ótica Visão",              38113.08,  "Negativou","Frio",   "abril",  "LeadBroker",        "ONE TIME",    "Bruno",       "",             ""),
    ("Faz Beauty",               23604.00,  "Proposta", "Frio",   "abril",  "LeadBroker",        "ONE TIME",    "Bruno",       "Bruno",        ""),
    ("Arteled Iluminação",       None,      "Proposta", "Frio",   "abril",  "LeadBroker",        "Ass. Booking","Bruno",       "Bruno",        ""),
    ("VegasOn",                  18000.00,  "Negativou","Frio",   "abril",  "Recuperação Broker","ONE TIME",    "João",        "",             ""),
    ("Buffet Primavera",         14050.00,  "Negativou","Frio",   "abril",  "LeadBroker",        "ONE TIME",    "Bruno",       "Daniel",       ""),
    ("MariasBobinas",            14050.00,  "Negativou","Frio",   "abril",  "LeadBroker",        "ONE TIME",    "Bruno",       "",             ""),
    ("Mendes Pereira e Taveira", 36000.00,  "Negativou","Frio",   "abril",  "LeadBroker",        "ONE TIME",    "João",        "",             ""),
    ("Mendes Pereira e Taveira", 4000.00,   "Negativou","Frio",   "abril",  "LeadBroker",        "Ass. Booking","João",        "",             ""),
    ("LuiOfalmologia",           18546.00,  "Negativou","Frio",   "Março",  "Recomendação",      "ONE TIME",    "Bruno",       "",             ""),
    ("Radar de Licitações",      3000.00,   "Negativou","Frio",   "abril",  "LeadBroker",        "ONE TIME",    "João",        "",             ""),
    ("Dr. Rafael",               3900.00,   "Negativou","Frio",   "abril",  "LeadBroker",        "Ass. Booking","Bruno",       "",             ""),
    ("Dra. Silvia Odete",        28100.00,  "Stand by", "Morno",  "Março",  "LeadBroker",        "ONE TIME",    "Bruno",       "",             ""),
    ("Tenext",                   5018.00,   "Stand by", "Frio",   "Março",  "Box",               "Ass. Booking","Bruno",       "Marcelo",      ""),
]

for ri, deal in enumerate(deals, 4):
    cliente, valor, status, temp, data_lead, canal, produto, closer, sdr, data_ass = deal
    bg_row = STATUS_BG.get(status, CARD)
    fg_temp = TEMP_FG.get(temp, WHITE)

    c(ws2, ri, 1, cliente,   bg=bg_row, bold=(status=="Fechado"), fg=WHITE if status!="Fechado" else ACCENT)
    c(ws2, ri, 2, valor or "?", bg=bg_row, fmt='R$ #.##0,00' if valor else None, align="right")
    c(ws2, ri, 3, status,    bg=STATUS_BG.get(status, CARD), fg=ACCENT if status=="Fechado" else WHITE, bold=True, align="center")
    c(ws2, ri, 4, temp,      bg=bg_row, fg=fg_temp, align="center")
    c(ws2, ri, 5, data_lead, bg=bg_row, fg=DIM, align="center")
    c(ws2, ri, 6, canal,     bg=bg_row, fg=DIM)
    c(ws2, ri, 7, produto,   bg=bg_row, fg=CYAN, align="center")
    c(ws2, ri, 8, closer,    bg=bg_row, fg=WHITE, align="center")
    c(ws2, ri, 9, sdr,       bg=bg_row, fg=DIM, align="center")
    c(ws2, ri, 10, data_ass, bg=bg_row, fg=ACCENT if data_ass else DIM, align="center")
    ws2.row_dimensions[ri].height = 18

# Linha de total
r_tot = len(deals) + 4
ws2.merge_cells(start_row=r_tot, start_column=1, end_row=r_tot, end_column=1)
c(ws2, r_tot, 1, f"TOTAL  ({len(deals)} deals)", bg=BLACK, fg=ACCENT, bold=True)
c(ws2, r_tot, 2, 502778.93, bg=BLACK, fg=ACCENT, bold=True, fmt='R$ #.##0,00', align="right")
for ci in range(3, 11):
    c(ws2, r_tot, ci, "", bg=BLACK)

# Dropdowns
for dv_formula, col in [
    ('"Fechado,Proposta,Negativou,Stand by"', 'C'),
    ('"Quente,Morno,Frio"', 'D'),
    ('"LeadBroker,Evento,Recuperação Broker,Recomendação,Box"', 'F'),
    ('"ONE TIME,Ass. Booking"', 'G'),
    ('"Bruno,João,Luis Felipe"', 'H'),
    ('"Daniel,Marcelo,João,Bruno,Luis Felipe"', 'I'),
]:
    dv = DataValidation(type="list", formula1=dv_formula)
    ws2.add_data_validation(dv)
    dv.sqref = f"{col}4:{col}100"

set_cols(ws2, [24, 13, 12, 12, 10, 20, 14, 14, 14, 14])
ws2.freeze_panes = "A4"


# ════════════════════════════════════════════════════════════
# ABA 3 — SDR DIÁRIO (Show / No Show / Marcadas / Remarcada)
# ════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("📞 SDR Diário")
ws3.sheet_properties.tabColor = CYAN

banner(ws3, 1, "CONTROLE SDR — ABRIL 2026", 7)

sdr_col_hdrs = ["DATA", "DIA", "SHOW", "NO SHOW", "MARCADAS", "REMARCADA", "OBS"]
for ci, h in enumerate(sdr_col_hdrs, 1):
    hdr(ws3, 2, ci, h)

# Dados do mês (do PDF — preenchidos até dia 23)
sdr_dias = [
    # (data, dia_sem, show, no_show, marcadas, remarcada)
    (date(2026,4,1),  "Qua", 8, 0, 0, 0),
    (date(2026,4,2),  "Qui", 0, 1, 4, 0),
    (date(2026,4,3),  "Sex", 0, 0, 0, 0),
    (date(2026,4,6),  "Seg", 3, 1, 5, 1),
    (date(2026,4,7),  "Ter", 3, 0, 1, 2),
    (date(2026,4,8),  "Qua", 2, 0, 1, 2),
    (date(2026,4,9),  "Qui", 3, 0, 4, 0),
    (date(2026,4,10), "Sex", 2, 1, 2, 0),
    (date(2026,4,13), "Seg", 1, 1, 4, 0),
    (date(2026,4,14), "Ter", 1, 0, 3, 0),
    (date(2026,4,15), "Qua", 2, 0, 6, 0),
    (date(2026,4,16), "Qui", 2, 1, 1, 0),
    (date(2026,4,17), "Sex", 1, 2, 3, 0),
    (date(2026,4,20), "Seg", 0, 1, 2, 2),
    (date(2026,4,21), "Ter", 1, 1, 4, 0),
    (date(2026,4,22), "Qua", 3, 2, None, None),
    (date(2026,4,23), "Qui", 1, None, None, None),
]
# Dias futuros em branco
dias_preenchidos = {d[0] for d in sdr_dias}
for dia in DIAS_UTEIS:
    if dia not in dias_preenchidos and dia > date(2026,4,23):
        sdr_dias.append((dia, ["Seg","Ter","Qua","Qui","Sex"][dia.weekday()], None, None, None, None))

sdr_dias.sort(key=lambda x: x[0])

for ri, (data, dia_sem, show, no_show, marcadas, remarcada) in enumerate(sdr_dias, 3):
    is_past = data <= date(2026,4,23)
    is_today = data == date(2026,4,23)
    row_bg = MID if is_past else CARD
    if is_today: row_bg = "1A2010"

    c(ws3, ri, 1, data.strftime("%d/%m/%Y"), bg=row_bg, fg=ACCENT if is_today else WHITE, bold=is_today, align="center")
    c(ws3, ri, 2, dia_sem, bg=row_bg, fg=DIM, align="center")
    c(ws3, ri, 3, show if show is not None else "", bg=G_BG if show else row_bg, fg=ACCENT, bold=True, align="center")
    c(ws3, ri, 4, no_show if no_show is not None else "", bg=R_BG if no_show else row_bg, fg="FF6666" if no_show else DIM, align="center")
    c(ws3, ri, 5, marcadas if marcadas is not None else "", bg=BLUE_BG if marcadas else row_bg, fg=CYAN if marcadas else DIM, bold=True, align="center")
    c(ws3, ri, 6, remarcada if remarcada is not None else "", bg=O_BG if remarcada else row_bg, fg="FF9900" if remarcada else DIM, align="center")
    c(ws3, ri, 7, "", bg=row_bg)
    ws3.row_dimensions[ri].height = 18

# Linha de totais
r_tot3 = len(sdr_dias) + 3
c(ws3, r_tot3, 1, "TOTAL", bg=BLACK, fg=ACCENT, bold=True, align="center")
c(ws3, r_tot3, 2, "", bg=BLACK)
c(ws3, r_tot3, 3, 33,  bg=G_BG,   fg=ACCENT, bold=True, align="center")
c(ws3, r_tot3, 4, 9,   bg=R_BG,   fg="FF6666", bold=True, align="center")
c(ws3, r_tot3, 5, 42,  bg=BLUE_BG,fg=CYAN, bold=True, align="center")
c(ws3, r_tot3, 6, 7,   bg=O_BG,   fg="FF9900", bold=True, align="center")
c(ws3, r_tot3, 7, "",  bg=BLACK)
ws3.row_dimensions[r_tot3].height = 22

set_cols(ws3, [13, 7, 9, 11, 11, 12, 20])
ws3.freeze_panes = "A3"


# ════════════════════════════════════════════════════════════
# ABA 4 — LEADBROKER DIÁRIO (leads comprados por dia)
# ════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("📡 Leadbroker")
ws4.sheet_properties.tabColor = "FF4444"

banner(ws4, 1, "LEADBROKER DIÁRIO — ABRIL 2026", 5, bg="1A0000", fg="FF4444")

# KPIs fixos no topo
lb_kpis = [
    ("Orçamento Mensal", "R$ 100.000,00"),
    ("Orçamento Semanal", "R$ 33.333,33"),
    ("Orçamento Diário", "R$ 5.000,00"),
    ("Ticket Médio Estimado", "R$ 900,00"),
    ("Leads Estimados/Mês", "~111"),
]
for ci, (lbl, val) in enumerate(lb_kpis, 1):
    c(ws4, 2, ci, lbl, bg=DARK, fg=DIM, sz=8, align="center")
    c(ws4, 3, ci, val, bg=MID, fg="FF4444", bold=True, sz=11, align="center")

ws4.row_dimensions[2].height = 14
ws4.row_dimensions[3].height = 22

lb_hdrs = ["DATA", "Nº LEADS", "VALOR TOTAL (R$)", "TICKET MÉDIO (R$)", "OBS"]
for ci, h in enumerate(lb_hdrs, 1):
    hdr(ws4, 4, ci, h, bg="1A0000", fg="FF4444")

for ri, dia in enumerate(DIAS_UTEIS, 5):
    is_past = dia <= date(2026,4,23)
    row_bg = MID if is_past else CARD
    c(ws4, ri, 1, dia.strftime("%d/%m/%Y"), bg=row_bg, align="center")
    for ci in range(2, 6):
        c(ws4, ri, ci, "", bg=G_BG if is_past else CARD)
    ws4.row_dimensions[ri].height = 18

# Total
r_tot4 = len(DIAS_UTEIS) + 5
c(ws4, r_tot4, 1, "TOTAL", bg=BLACK, fg="FF4444", bold=True, align="center")
for ci in range(2, 6):
    c(ws4, r_tot4, ci, "", bg=BLACK)

set_cols(ws4, [13, 10, 18, 18, 24])
ws4.freeze_panes = "A5"


# ════════════════════════════════════════════════════════════
# ABA 5 — METAS SEMANAIS (pacing por semana)
# ════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("📆 Pacing Semanal")
ws5.sheet_properties.tabColor = "888888"

banner(ws5, 1, "PACING SEMANAL — ABRIL 2026", 5)

semanas = [
    ("Semana 1", "01/04 – 04/04", 52500.00,  52500.00),
    ("Semana 2", "06/04 – 11/04", 70000.00,  70000.00),
    ("Semana 3", "13/04 – 18/04", 87500.00,  87500.00),
    ("Semana 4", "20/04 – 30/04", 140000.00, 140000.00),
]

sem_hdrs = ["SEMANA", "PERÍODO", "META JOÃO", "META BRUNO", "RECEITA REAL"]
for ci, h in enumerate(sem_hdrs, 1):
    hdr(ws5, 2, ci, h)

for ri, (sem, per, meta_j, meta_b) in enumerate(semanas, 3):
    c(ws5, ri, 1, sem, bg=DARK, fg=ACCENT, bold=True, align="center")
    c(ws5, ri, 2, per, bg=CARD, fg=DIM, align="center")
    c(ws5, ri, 3, meta_j, bg=CARD, fmt='R$ #.##0,00', align="right")
    c(ws5, ri, 4, meta_b, bg=CARD, fmt='R$ #.##0,00', align="right")
    c(ws5, ri, 5, "", bg=G_BG)  # preencher
    ws5.row_dimensions[ri].height = 22

r5 = 7
section_row(ws5, r5, "GAPS ACUMULADOS (João vs Bruno)", 5)
r5 += 1
gap_hdrs = ["SEMANA", "GAP JOÃO", "GAP BRUNO", "GAP TOTAL", ""]
gap_vals = [
    ("S1", -18975.00, -19361.09, None, ""),
    ("S2", -36475.00, -36861.09, None, ""),
    ("S3", -53975.00, -54361.09, None, ""),
    ("S4", -106475.00, -217444.36, None, ""),
]
for ci, h in enumerate(gap_hdrs, 1):
    hdr(ws5, r5, ci, h)
r5 += 1
for row_d in gap_vals:
    sem, gj, gb, gt, _ = row_d
    c(ws5, r5, 1, sem, bg=DARK, fg=ACCENT, bold=True, align="center")
    c(ws5, r5, 2, gj, bg=R_BG, fg="FF6666", fmt='R$ #.##0,00', align="right")
    c(ws5, r5, 3, gb, bg=R_BG, fg="FF6666", fmt='R$ #.##0,00', align="right")
    c(ws5, r5, 4, "", bg=CARD)
    c(ws5, r5, 5, "", bg=CARD)
    r5 += 1

set_cols(ws5, [12, 18, 16, 16, 16])
ws5.freeze_panes = "A3"


# ════════════════════════════════════════════════════════════
# Salvar
# ════════════════════════════════════════════════════════════
out = "/home/oxicoreco/Documentos/GitHub/oxicorerevenue-main/docs/oxicore_abr2026.xlsx"
wb.save(out)
print(f"Salvo: {out}")
print(f"Deals: {len(deals)}")
print(f"Dias úteis: {len(DIAS_UTEIS)}")
