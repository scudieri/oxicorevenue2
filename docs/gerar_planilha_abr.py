import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import date, timedelta

wb = openpyxl.Workbook()

# ── Paleta ───────────────────────────────────────────────────────────────
BLACK     = "0D0D0D"
DARK      = "1A1A1A"
MID       = "252525"
CARD      = "2A2A2A"
ACCENT    = "C8FF00"   # verde limão
CYAN      = "00E5FF"
WHITE     = "FFFFFF"
DIM       = "777777"
GREEN_BG  = "1A3320"
RED_BG    = "3A1A1A"
ORANGE_BG = "2A2010"

def fill(c): return PatternFill("solid", fgColor=c)
def fnt(bold=False, color=WHITE, size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")
def bdr(color="333333"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)
def aln(h="left", wrap=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

def hdr(ws, r, c, text, bg=DARK, fg=ACCENT, sz=9, bold=True):
    cell = ws.cell(r, c, text)
    cell.fill = fill(bg); cell.font = fnt(bold, fg, sz)
    cell.alignment = aln("center"); cell.border = bdr()
    return cell

def cell(ws, r, c, v="", bg=CARD, fg=WHITE, bold=False, fmt=None, align="left"):
    cl = ws.cell(r, c, v)
    cl.fill = fill(bg); cl.font = fnt(bold, fg)
    cl.alignment = aln(align); cl.border = bdr()
    if fmt: cl.number_format = fmt
    return cl

def hint(ws, r, c, text):
    cl = ws.cell(r, c, text)
    cl.fill = fill("181818"); cl.font = fnt(False, DIM, 8, True)
    cl.alignment = aln("center", wrap=True); cl.border = bdr()

def banner(ws, r, text, ncols, bg=BLACK, fg=ACCENT, sz=13):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    cl = ws.cell(r, 1, text)
    cl.fill = fill(bg); cl.font = fnt(True, fg, sz)
    cl.alignment = aln("center"); ws.row_dimensions[r].height = 30

def sub(ws, r, text, ncols):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    cl = ws.cell(r, 1, text)
    cl.fill = fill("131313"); cl.font = fnt(False, DIM, 8, True)
    cl.alignment = aln("center"); ws.row_dimensions[r].height = 14

def set_cols(ws, widths):
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# Dias úteis ABR 2026 (sem Tiradentes 21/04)
DIAS_UTEIS = []
d = date(2026, 4, 1)
feriados = {date(2026, 4, 21)}
while d.month == 4:
    if d.weekday() < 5 and d not in feriados:
        DIAS_UTEIS.append(d)
    d += timedelta(1)


# ══════════════════════════════════════════════════════════════
# ABA 1 — CONFIG ABR 2026
# ══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "⚙️ Config ABR 2026"
ws1.sheet_properties.tabColor = ACCENT

NC = 3
banner(ws1, 1, "CONFIG — ABRIL 2026", NC)
sub(ws1, 2, "Preencha apenas os valores em verde antes de começar o mês.", NC)

# Bloco de metas
rows_config = [
    ("── METAS ──",                  "",         ""),
    ("Meta ONETIME (R$)",            180000,     "vendas pontuais / projetos"),
    ("Meta MRR — Recorrente (R$)",   55000,      "vendas de contrato mensal"),
    ("Meta Total (R$)",              235000,     "= ONETIME + MRR"),
    ("",                             "",         ""),
    ("── CANAL DE AQUISIÇÃO ──",     "",         ""),
    ("Canal",                        "Leadbroker",""),
    ("Investimento (R$)",            100000,     "verba total do mês"),
    ("",                             "",         ""),
    ("── EQUIPE ──",                 "",         ""),
    ("Closers",                      "Bruno, João",     "nomes separados por vírgula"),
    ("SDRs",                         "Daniel, Marcelo", "nomes separados por vírgula"),
    ("",                             "",         ""),
    ("Dias Úteis do Mês",            21,         "calculado automaticamente (sem Tiradentes)"),
]

for i, (label, value, note) in enumerate(rows_config, 4):
    if label == "":
        ws1.row_dimensions[i].height = 8
        continue
    is_section = label.startswith("──")
    if is_section:
        ws1.merge_cells(start_row=i, start_column=1, end_row=i, end_column=3)
        cl = ws1.cell(i, 1, label.replace("──", "").strip())
        cl.fill = fill(BLACK); cl.font = fnt(True, CYAN, 9)
        cl.alignment = aln("center"); cl.border = bdr()
        ws1.row_dimensions[i].height = 16
        continue

    cl = ws1.cell(i, 1, label)
    cl.fill = fill(DARK); cl.font = fnt(True, DIM, 9)
    cl.alignment = aln(); cl.border = bdr()

    is_fixed = value != ""
    bg = MID if is_fixed else GREEN_BG
    fg = ACCENT if is_fixed else ACCENT
    cl2 = ws1.cell(i, 2, value)
    cl2.fill = fill(bg); cl2.font = fnt(True, fg, 11)
    cl2.alignment = aln("center"); cl2.border = bdr()
    if isinstance(value, (int, float)) and "R$" in label or "Investimento" in label or "Meta" in label:
        cl2.number_format = 'R$ #.##0,00'

    cl3 = ws1.cell(i, 3, note)
    cl3.fill = fill("181818"); cl3.font = fnt(False, DIM, 8, True)
    cl3.alignment = aln("left"); cl3.border = bdr()
    ws1.row_dimensions[i].height = 22

# Bloco de dias úteis
r = len(rows_config) + 4 + 2
ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
cl = ws1.cell(r, 1, f"📆  DIAS ÚTEIS DE ABRIL 2026  —  {len(DIAS_UTEIS)} dias  (Tiradentes 21/04 excluído)")
cl.fill = fill(BLACK); cl.font = fnt(True, CYAN, 10)
cl.alignment = aln("center"); ws1.row_dimensions[r].height = 22
r += 1

dias_por_linha = 7
for i, dia in enumerate(DIAS_UTEIS):
    col = (i % dias_por_linha) + 1
    row = r + (i // dias_por_linha)
    weekday = ["Seg","Ter","Qua","Qui","Sex"][dia.weekday()]
    cl = ws1.cell(row, col, f"{dia.day:02d}/04  {weekday}")
    cl.fill = fill(GREEN_BG); cl.font = fnt(False, ACCENT, 9)
    cl.alignment = aln("center"); cl.border = bdr()
    ws1.row_dimensions[row].height = 18

set_cols(ws1, [28, 22, 32])
ws1.freeze_panes = "A4"


# ══════════════════════════════════════════════════════════════
# ABA 2 — LEADS
# ══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("🔥 Leads")
ws2.sheet_properties.tabColor = ACCENT

NC2 = 4
banner(ws2, 1, "LEADS — ABRIL 2026", NC2, fg=BLACK, bg=ACCENT)
sub(ws2, 2, "Registre diariamente. Pode consolidar por origem (não precisa ser lead por lead).", NC2)

cols2 = ["data",        "quantidade",  "origem",                    "observacao"]
hints2 = ["2026-04-01", "ex: 47",      "Instagram / Google / etc.", "opcional"]
for c, (h, hn) in enumerate(zip(cols2, hints2), 1):
    hdr(ws2, 3, c, h.upper(), bg=DARK, fg=ACCENT)
    hint(ws2, 4, c, hn)

# Pré-preenche os dias úteis na coluna data
for i, dia in enumerate(DIAS_UTEIS, 5):
    cell(ws2, i, 1, dia.strftime("%Y-%m-%d"), bg=MID)
    cell(ws2, i, 2, "", bg=GREEN_BG)   # quantidade — preencher
    cell(ws2, i, 3, "", bg=GREEN_BG)   # origem — preencher
    cell(ws2, i, 4, "", bg=CARD)

dv_orig = DataValidation(type="list",
    formula1='"Instagram,Google Ads,Meta Ads,LinkedIn,TikTok,Indicação,Outros"')
ws2.add_data_validation(dv_orig)
dv_orig.sqref = f"C5:C{4+len(DIAS_UTEIS)+20}"

set_cols(ws2, [14, 13, 18, 28])
ws2.freeze_panes = "A5"


# ══════════════════════════════════════════════════════════════
# ABA 3 — REUNIÕES
# ══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("📞 Reuniões")
ws3.sheet_properties.tabColor = CYAN

NC3 = 7
banner(ws3, 1, "REUNIÕES — ABRIL 2026", NC3)
sub(ws3, 2, "SDR cria a linha ao marcar. Atualiza o status após a reunião.", NC3)

cols3 = ["sdr",          "closer",      "data_agendamento", "data_reuniao", "status",                          "cliente / obs",      "remarcada para"]
hints3 = ["Daniel/Marcelo","Bruno/João", "data que marcou",  "data prevista","marcada → show/no_show/remarcada", "nome ou observação", "se remarcada"]
for c, (h, hn) in enumerate(zip(cols3, hints3), 1):
    hdr(ws3, 3, c, h.upper())
    hint(ws3, 4, c, hn)

# Linhas em branco para preencher
for r in range(5, 105):
    for c in range(1, 8):
        bg = GREEN_BG if c == 5 else CARD
        cell(ws3, r, c, bg=bg)

dv_status = DataValidation(type="list",
    formula1='"marcada,show,no_show,remarcada,cancelada"')
ws3.add_data_validation(dv_status)
dv_status.sqref = "E5:E104"

dv_sdr = DataValidation(type="list", formula1='"Daniel,Marcelo"')
ws3.add_data_validation(dv_sdr)
dv_sdr.sqref = "A5:A104"

dv_closer = DataValidation(type="list", formula1='"Bruno,João"')
ws3.add_data_validation(dv_closer)
dv_closer.sqref = "B5:B104"

# Exemplo linha 5
exemplo_r = [("Daniel","Bruno","2026-04-01","2026-04-03","show","Empresa ABC Ltda","")]
for c, v in enumerate(exemplo_r[0], 1):
    bg = {"show": GREEN_BG, "no_show": RED_BG, "marcada": CARD}.get(v, CARD)
    cell(ws3, 5, c, v, bg=bg)

set_cols(ws3, [12, 12, 16, 14, 14, 26, 14])
ws3.freeze_panes = "A5"


# ══════════════════════════════════════════════════════════════
# ABA 4 — VENDAS
# ══════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("💎 Vendas")
ws4.sheet_properties.tabColor = "FF9900"

NC4 = 7
banner(ws4, 1, "VENDAS — ABRIL 2026", NC4, bg="1A1200", fg="FF9900")
sub(ws4, 2, "Closer registra ao fechar. Tipo: 'onetime' = venda única  |  'recorrente' = contrato MRR", NC4)

cols4 = ["data_venda",  "closer",    "tipo",                  "valor (R$)", "status",                    "cliente_nome",   "observacao"]
hints4 = ["2026-04-07", "Bruno/João","onetime / recorrente",  "ex: 14900",  "confirmado/na_rua/cancelado","Empresa XPTO",   "opcional"]
for c, (h, hn) in enumerate(zip(cols4, hints4), 1):
    hdr(ws4, 3, c, h.upper(), bg="1A1200", fg="FF9900")
    hint(ws4, 4, c, hn)

for r in range(5, 65):
    for c in range(1, 8):
        bg = GREEN_BG if c == 3 else CARD
        cell(ws4, r, c, bg=bg)

# Exemplo
ex4 = [("2026-04-07", "Bruno", "onetime", 14900, "confirmado", "Empresa XPTO", "Pagamento à vista")]
for c, v in enumerate(ex4[0], 1):
    bg = {"confirmado": GREEN_BG, "na_rua": ORANGE_BG, "cancelado": RED_BG,
          "onetime": "1A2030", "recorrente": "1A3020"}.get(str(v), CARD)
    fmt = 'R$ #.##0,00' if c == 4 else None
    cell(ws4, 5, c, v, bg=bg, fmt=fmt)

dv_vtipo = DataValidation(type="list", formula1='"onetime,recorrente"')
ws4.add_data_validation(dv_vtipo)
dv_vtipo.sqref = "C5:C64"

dv_vstatus = DataValidation(type="list", formula1='"confirmado,na_rua,cancelado"')
ws4.add_data_validation(dv_vstatus)
dv_vstatus.sqref = "E5:E64"

dv_vcloser = DataValidation(type="list", formula1='"Bruno,João"')
ws4.add_data_validation(dv_vcloser)
dv_vcloser.sqref = "B5:B64"

set_cols(ws4, [14, 12, 14, 14, 16, 24, 28])
ws4.freeze_panes = "A5"


# ══════════════════════════════════════════════════════════════
# ABA 5 — RESUMO (fórmulas automáticas)
# ══════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("📊 Resumo")
ws5.sheet_properties.tabColor = "555555"

banner(ws5, 1, "RESUMO ABR 2026 — calculado automaticamente", 4)
sub(ws5, 2, "Não edite esta aba. Os valores atualizam conforme você preenche as outras abas.", 4)

def kpi_row(ws, r, label, formula, fmt=None):
    cl = ws.cell(r, 1, label)
    cl.fill = fill(DARK); cl.font = fnt(True, DIM, 9)
    cl.alignment = aln(); cl.border = bdr()
    ws.row_dimensions[r].height = 22

    cl2 = ws.cell(r, 2, formula)
    cl2.fill = fill(MID); cl2.font = fnt(True, ACCENT, 12)
    cl2.alignment = aln("center"); cl2.border = bdr()
    if fmt: cl2.number_format = fmt

    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    cl3 = ws.cell(r, 3, "")
    cl3.fill = fill(BLACK); cl3.border = bdr()

# Nota: Vendas — col B=closer, C=tipo, D=valor, E=status
V = "'💎 Vendas'"
R = "'📞 Reuniões'"
L = "'🔥 Leads'"
LEADS_RANGE = f"{L}!A5:A{4+len(DIAS_UTEIS)}"

sections = [
    ("── FUNIL (Leadbroker) ──", None, None),
    ("Total de Leads",             f"=SUMIF({L}!B5:B200,\">0\",{L}!B5:B200)", None),
    ("Reuniões Marcadas",          f"=COUNTA({R}!A5:A200)-COUNTBLANK({R}!A5:A200)", None),
    ("Shows (acontecidas)",        f"=COUNTIF({R}!E5:E200,\"show\")", None),
    ("No-Shows",                   f"=COUNTIF({R}!E5:E200,\"no_show\")", None),
    ("Remarcadas",                 f"=COUNTIF({R}!E5:E200,\"remarcada\")", None),
    ("── ONETIME — Meta R$ 180K ──", None, None),
    ("Vendas Onetime Confirmadas", f"=COUNTIFS({V}!C5:C64,\"onetime\",{V}!E5:E64,\"confirmado\")", None),
    ("Receita Onetime Confirmada", f"=SUMIFS({V}!D5:D64,{V}!C5:C64,\"onetime\",{V}!E5:E64,\"confirmado\")", "R$ #.##0,00"),
    ("Onetime Na Rua",             f"=SUMIFS({V}!D5:D64,{V}!C5:C64,\"onetime\",{V}!E5:E64,\"na_rua\")", "R$ #.##0,00"),
    ("% da Meta Onetime",          f"=IFERROR(SUMIFS({V}!D5:D64,{V}!C5:C64,\"onetime\",{V}!E5:E64,\"confirmado\")/180000,0)", "0,0%"),
    ("── MRR — Meta R$ 55K ──", None, None),
    ("Contratos MRR Confirmados",  f"=COUNTIFS({V}!C5:C64,\"recorrente\",{V}!E5:E64,\"confirmado\")", None),
    ("Receita MRR Confirmada",     f"=SUMIFS({V}!D5:D64,{V}!C5:C64,\"recorrente\",{V}!E5:E64,\"confirmado\")", "R$ #.##0,00"),
    ("MRR Na Rua",                 f"=SUMIFS({V}!D5:D64,{V}!C5:C64,\"recorrente\",{V}!E5:E64,\"na_rua\")", "R$ #.##0,00"),
    ("% da Meta MRR",              f"=IFERROR(SUMIFS({V}!D5:D64,{V}!C5:C64,\"recorrente\",{V}!E5:E64,\"confirmado\")/55000,0)", "0,0%"),
    ("── TOTAL ──", None, None),
    ("Receita Total Confirmada",   f"=SUMIF({V}!E5:E64,\"confirmado\",{V}!D5:D64)", "R$ #.##0,00"),
    ("Total Na Rua",               f"=SUMIF({V}!E5:E64,\"na_rua\",{V}!D5:D64)", "R$ #.##0,00"),
    ("% da Meta Total (R$ 235K)",  f"=IFERROR(SUMIF({V}!E5:E64,\"confirmado\",{V}!D5:D64)/235000,0)", "0,0%"),
    ("── CLOSERS ──", None, None),
    ("Bruno — Receita",            f"=SUMIFS({V}!D5:D64,{V}!B5:B64,\"Bruno\",{V}!E5:E64,\"confirmado\")", "R$ #.##0,00"),
    ("João — Receita",             f"=SUMIFS({V}!D5:D64,{V}!B5:B64,\"João\",{V}!E5:E64,\"confirmado\")", "R$ #.##0,00"),
    ("── SDRs ──", None, None),
    ("Daniel — Shows",             f"=COUNTIFS({R}!A5:A200,\"Daniel\",{R}!E5:E200,\"show\")", None),
    ("Marcelo — Shows",            f"=COUNTIFS({R}!A5:A200,\"Marcelo\",{R}!E5:E200,\"show\")", None),
    ("── PACING (vs Meta Total) ──", None, None),
    ("Dias Úteis do Mês",          "=21", None),
    ("Dias Úteis Passados",        f"=COUNTIF({LEADS_RANGE},\"<=\"&TODAY())", None),
    ("Receita Ideal Hoje",         f"=IFERROR(235000/21*COUNTIF({LEADS_RANGE},\"<=\"&TODAY()),0)", "R$ #.##0,00"),
    ("Gap (Real - Ideal)",         f"=IFERROR(SUMIF({V}!E5:E64,\"confirmado\",{V}!D5:D64)-(235000/21*COUNTIF({LEADS_RANGE},\"<=\"&TODAY())),0)", "R$ #.##0,00"),
]

r = 3
for label, formula, fmt in sections:
    if formula is None:
        ws5.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        cl = ws5.cell(r, 1, label)
        cl.fill = fill(BLACK); cl.font = fnt(True, CYAN, 9)
        cl.alignment = aln("center"); cl.border = bdr()
        ws5.row_dimensions[r].height = 14
    else:
        kpi_row(ws5, r, label, formula, fmt)
    r += 1

set_cols(ws5, [26, 20, 10, 10])


# ══════════════════════════════════════════════════════════════
# Salvar
# ══════════════════════════════════════════════════════════════
out = "/home/oxicoreco/Documentos/GitHub/oxicorerevenue-main/docs/oxicore_abr2026.xlsx"
wb.save(out)
print(f"Salvo: {out}")
