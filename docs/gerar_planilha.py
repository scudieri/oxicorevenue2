import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()

# ── Paleta ──────────────────────────────────────────────────────────────
BLACK      = "0D0D0D"
DARK_GRAY  = "1A1A1A"
MID_GRAY   = "2A2A2A"
LIGHT_GRAY = "3A3A3A"
ACCENT     = "C8FF00"       # verde limão (identidade do projeto)
ACCENT2    = "00E5FF"       # ciano
WHITE      = "FFFFFF"
TEXT_DIM   = "888888"
RED        = "FF4444"
ORANGE     = "FF9900"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=WHITE, size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Calibri")

def border_thin(color="333333"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left(wrap=False):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def apply_header(ws, row, col, text,
                 bg=DARK_GRAY, fg=ACCENT, sz=10, bold=True, merged_to=None):
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill = fill(bg)
    cell.font = font(bold=bold, color=fg, size=sz)
    cell.alignment = center(wrap=True)
    cell.border = border_thin()
    return cell

def apply_cell(ws, row, col, value="",
               bg=MID_GRAY, fg=WHITE, bold=False,
               align="left", number_fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = fill(bg)
    cell.font = font(bold=bold, color=fg)
    cell.alignment = center() if align == "center" else left(wrap=True)
    cell.border = border_thin()
    if number_fmt:
        cell.number_format = number_fmt
    return cell

def note_cell(ws, row, col, text):
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill = fill("1E1E1E")
    cell.font = font(color=TEXT_DIM, size=8, italic=True)
    cell.alignment = left(wrap=True)
    return cell

def title_row(ws, row, text, ncols):
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = fill(BLACK)
    cell.font = font(bold=True, color=ACCENT, size=13)
    cell.alignment = center()
    ws.row_dimensions[row].height = 28

def subtitle_row(ws, row, text, ncols):
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = fill("181818")
    cell.font = font(bold=False, color=TEXT_DIM, size=9, italic=True)
    cell.alignment = center()
    ws.row_dimensions[row].height = 16

def freeze(ws, cell="A3"):
    ws.freeze_panes = cell

def tab_color(ws, hex_color):
    ws.sheet_properties.tabColor = hex_color


# ════════════════════════════════════════════════════════════════════════
# ABA 1 — PERÍODOS
# ════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "📅 Períodos"
tab_color(ws, ACCENT)

NCOLS = 6
title_row(ws, 1, "PERÍODOS — Criar um registro por mês de operação", NCOLS)
subtitle_row(ws, 2, "Preencha no início de cada mês. O campo 'ativo' deve ser SIM apenas para o mês vigente.", NCOLS)

headers = ["nome", "mes (1-12)", "ano", "data_inicio", "data_fim", "ativo"]
hints   = ["ex: FEV 2026", "ex: 2", "ex: 2026", "ex: 2026-02-01", "ex: 2026-02-28", "SIM / NÃO"]

for c, (h, hint) in enumerate(zip(headers, hints), 1):
    apply_header(ws, 3, c, h)
    note_cell(ws, 4, c, hint)

# Linha exemplo
example = ["FEV 2026", 2, 2026, "2026-02-01", "2026-02-28", "SIM"]
for c, v in enumerate(example, 1):
    apply_cell(ws, 5, c, v, bg=LIGHT_GRAY)

# Validação SIM/NÃO
dv = DataValidation(type="list", formula1='"SIM,NÃO"', showDropDown=False)
ws.add_data_validation(dv)
dv.sqref = "F5:F200"

ws.column_dimensions["A"].width = 14
ws.column_dimensions["B"].width = 12
ws.column_dimensions["C"].width = 8
ws.column_dimensions["D"].width = 14
ws.column_dimensions["E"].width = 14
ws.column_dimensions["F"].width = 10
ws.row_dimensions[3].height = 22
ws.row_dimensions[4].height = 14
freeze(ws, "A5")


# ════════════════════════════════════════════════════════════════════════
# ABA 2 — PESSOAS
# ════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("👤 Pessoas")
tab_color(ws2, "00E5FF")

NCOLS2 = 4
title_row(ws2, 1, "PESSOAS — Cadastro de Closers e SDRs", NCOLS2)
subtitle_row(ws2, 2, "Cadastre uma vez. Papel define se é closer, sdr ou ambos.", NCOLS2)

headers2 = ["nome", "papel", "avatar_url", "ativo"]
hints2   = ["ex: Bruno", "closer / sdr / ambos", "URL da foto (opcional)", "SIM / NÃO"]

for c, (h, hint) in enumerate(zip(headers2, hints2), 1):
    apply_header(ws2, 3, c, h)
    note_cell(ws2, 4, c, hint)

pessoas_exemplo = [
    ["Bruno",   "closer", "", "SIM"],
    ["João",    "closer", "", "SIM"],
    ["Daniel",  "sdr",    "", "SIM"],
    ["Marcelo", "sdr",    "", "SIM"],
]
for r, row in enumerate(pessoas_exemplo, 5):
    for c, v in enumerate(row, 1):
        apply_cell(ws2, r, c, v, bg=LIGHT_GRAY)

dv2 = DataValidation(type="list", formula1='"closer,sdr,ambos"')
ws2.add_data_validation(dv2)
dv2.sqref = "B5:B200"

dv3 = DataValidation(type="list", formula1='"SIM,NÃO"')
ws2.add_data_validation(dv3)
dv3.sqref = "D5:D200"

ws2.column_dimensions["A"].width = 16
ws2.column_dimensions["B"].width = 12
ws2.column_dimensions["C"].width = 30
ws2.column_dimensions["D"].width = 10
ws2.row_dimensions[3].height = 22
ws2.row_dimensions[4].height = 14
freeze(ws2, "A5")


# ════════════════════════════════════════════════════════════════════════
# ABA 3 — METAS MENSAIS
# ════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("🎯 Metas Mensais")
tab_color(ws3, "FF9900")

NCOLS3 = 6
title_row(ws3, 1, "METAS MENSAIS — Definir no início de cada mês", NCOLS3)
subtitle_row(ws3, 2, "Uma linha por período. Preencha antes de começar o mês.", NCOLS3)

headers3 = ["periodo", "meta_receita (R$)", "dias_uteis_mes", "meta_reunioes", "meta_vendas", "ticket_meta (R$)"]
hints3   = ["ex: FEV 2026", "ex: 150000", "ex: 20", "ex: 60", "ex: 12", "ex: 12500"]

for c, (h, hint) in enumerate(zip(headers3, hints3), 1):
    apply_header(ws3, 3, c, h)
    note_cell(ws3, 4, c, hint)

ex3 = ["FEV 2026", 150000, 20, 60, 12, 12500]
for c, v in enumerate(ex3, 1):
    fmt = 'R$ #.##0,00' if c in (2, 6) else None
    apply_cell(ws3, 5, c, v, bg=LIGHT_GRAY, number_fmt=fmt)

for col, w in zip("ABCDEF", [14, 18, 16, 14, 14, 18]):
    ws3.column_dimensions[col].width = w
ws3.row_dimensions[3].height = 22
ws3.row_dimensions[4].height = 14
freeze(ws3, "A5")


# ════════════════════════════════════════════════════════════════════════
# ABA 4 — DIAS ÚTEIS
# ════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("📆 Dias Úteis")
tab_color(ws4, "888888")

NCOLS4 = 2
title_row(ws4, 1, "DIAS ÚTEIS TRABALHADOS — Registrar todo dia útil", NCOLS4)
subtitle_row(ws4, 2, "Cada linha = um dia útil trabalhado. Usado para calcular o pacing.", NCOLS4)

headers4 = ["periodo", "data"]
hints4   = ["ex: FEV 2026", "ex: 2026-02-03"]

for c, (h, hint) in enumerate(zip(headers4, hints4), 1):
    apply_header(ws4, 3, c, h)
    note_cell(ws4, 4, c, hint)

dias_ex = [
    ["FEV 2026", "2026-02-03"],
    ["FEV 2026", "2026-02-04"],
    ["FEV 2026", "2026-02-05"],
]
for r, row in enumerate(dias_ex, 5):
    for c, v in enumerate(row, 1):
        apply_cell(ws4, r, c, v, bg=LIGHT_GRAY)

ws4.column_dimensions["A"].width = 14
ws4.column_dimensions["B"].width = 16
ws4.row_dimensions[3].height = 22
ws4.row_dimensions[4].height = 14
freeze(ws4, "A5")


# ════════════════════════════════════════════════════════════════════════
# ABA 5 — INVESTIMENTO MARKETING
# ════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("💰 Investimento")
tab_color(ws5, "FF4444")

NCOLS5 = 5
title_row(ws5, 1, "INVESTIMENTO DE MARKETING — Lançar conforme verba é aplicada", NCOLS5)
subtitle_row(ws5, 2, "Pode lançar o total do mês de uma vez, ou parcelar por campanha.", NCOLS5)

headers5 = ["periodo", "data_lancamento", "valor (R$)", "canal", "descricao"]
hints5   = ["ex: FEV 2026", "ex: 2026-02-01", "ex: 15000", "ex: Meta Ads", "ex: Campanha Topo de Funil"]

for c, (h, hint) in enumerate(zip(headers5, hints5), 1):
    apply_header(ws5, 3, c, h)
    note_cell(ws5, 4, c, hint)

ex5 = ["FEV 2026", "2026-02-01", 15000, "Meta Ads", "Campanha Topo de Funil"]
for c, v in enumerate(ex5, 1):
    fmt = 'R$ #.##0,00' if c == 3 else None
    apply_cell(ws5, 5, c, v, bg=LIGHT_GRAY, number_fmt=fmt)

dv5 = DataValidation(type="list", formula1='"Meta Ads,Google Ads,LinkedIn Ads,TikTok Ads,Outros"')
ws5.add_data_validation(dv5)
dv5.sqref = "D5:D200"

for col, w in zip("ABCDE", [14, 16, 14, 14, 32]):
    ws5.column_dimensions[col].width = w
ws5.row_dimensions[3].height = 22
ws5.row_dimensions[4].height = 14
freeze(ws5, "A5")


# ════════════════════════════════════════════════════════════════════════
# ABA 6 — LEADS
# ════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("🔥 Leads")
tab_color(ws6, "C8FF00")

NCOLS6 = 4
title_row(ws6, 1, "LEADS — Registrar entrada de leads por dia/origem", NCOLS6)
subtitle_row(ws6, 2, "Pode consolidar por dia (ex: 47 leads do Instagram em 03/02). Não precisa ser um por um.", NCOLS6)

headers6 = ["periodo", "data_entrada", "quantidade", "origem"]
hints6   = ["ex: FEV 2026", "ex: 2026-02-03", "ex: 47", "ex: Instagram"]

for c, (h, hint) in enumerate(zip(headers6, hints6), 1):
    apply_header(ws6, 3, c, h, fg=BLACK)
    note_cell(ws6, 4, c, hint)

ex6 = [
    ["FEV 2026", "2026-02-03", 47, "Instagram"],
    ["FEV 2026", "2026-02-03", 12, "Google"],
    ["FEV 2026", "2026-02-04", 35, "Instagram"],
]
for r, row in enumerate(ex6, 5):
    for c, v in enumerate(row, 1):
        apply_cell(ws6, r, c, v, bg=LIGHT_GRAY)

dv6 = DataValidation(type="list", formula1='"Instagram,Google,LinkedIn,TikTok,Indicação,Outros"')
ws6.add_data_validation(dv6)
dv6.sqref = "D5:D200"

for col, w in zip("ABCD", [14, 14, 12, 16]):
    ws6.column_dimensions[col].width = w
ws6.row_dimensions[3].height = 22
ws6.row_dimensions[4].height = 14
freeze(ws6, "A5")


# ════════════════════════════════════════════════════════════════════════
# ABA 7 — REUNIÕES
# ════════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("📞 Reuniões")
tab_color(ws7, "00E5FF")

NCOLS7 = 8
title_row(ws7, 1, "REUNIÕES — Registrar cada reunião marcada e atualizar o status", NCOLS7)
subtitle_row(ws7, 2, "SDR cria a linha ao marcar. Atualiza o status após a reunião acontecer (show/no_show).", NCOLS7)

headers7 = ["periodo", "sdr", "closer", "data_agendamento", "data_reuniao", "status", "observacao", "id_reuniao"]
hints7   = [
    "ex: FEV 2026", "ex: Daniel", "ex: Bruno",
    "ex: 2026-02-03", "ex: 2026-02-05",
    "marcada/show/no_show/remarcada/cancelada",
    "ex: CEO de startup EdTech", "gerado pelo sistema"
]

for c, (h, hint) in enumerate(zip(headers7, hints7), 1):
    apply_header(ws7, 3, c, h)
    note_cell(ws7, 4, c, hint)

ex7 = [
    ["FEV 2026", "Daniel",  "Bruno", "2026-02-03", "2026-02-05", "show",    "Decisor confirmado", 1],
    ["FEV 2026", "Daniel",  "João",  "2026-02-03", "2026-02-05", "no_show", "",                   2],
    ["FEV 2026", "Marcelo", "Bruno", "2026-02-04", "2026-02-06", "marcada", "Falar com sócio",    3],
]
for r, row in enumerate(ex7, 5):
    for c, v in enumerate(row, 1):
        bg = LIGHT_GRAY
        if c == 6:  # status col colorido
            bg = {"show": "1A3320", "no_show": "3A1A1A", "marcada": MID_GRAY,
                  "remarcada": "2A2010", "cancelada": "2A1A1A"}.get(v, MID_GRAY)
        apply_cell(ws7, r, c, v, bg=bg)

dv7 = DataValidation(type="list", formula1='"marcada,show,no_show,remarcada,cancelada"')
ws7.add_data_validation(dv7)
dv7.sqref = "F5:F500"

for col, w in zip("ABCDEFGH", [14, 12, 12, 16, 14, 16, 28, 12]):
    ws7.column_dimensions[col].width = w
ws7.row_dimensions[3].height = 22
ws7.row_dimensions[4].height = 14
freeze(ws7, "A5")


# ════════════════════════════════════════════════════════════════════════
# ABA 8 — VENDAS
# ════════════════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("💎 Vendas")
tab_color(ws8, "FF9900")

NCOLS8 = 8
title_row(ws8, 1, "VENDAS — Registrar cada fechamento", NCOLS8)
subtitle_row(ws8, 2, "Closer registra ao fechar. Status 'na_rua' = proposta aceita aguardando pgto. 'confirmado' = dinheiro na conta.", NCOLS8)

headers8 = ["periodo", "closer", "data_venda", "valor (R$)", "status", "cliente_nome", "id_reuniao (opt.)", "observacao"]
hints8   = [
    "ex: FEV 2026", "ex: Bruno", "ex: 2026-02-07",
    "ex: 14900", "confirmado/na_rua/cancelado",
    "ex: Empresa XPTO", "ex: 3 (da aba Reuniões)", "observação livre"
]

for c, (h, hint) in enumerate(zip(headers8, hints8), 1):
    apply_header(ws8, 3, c, h, fg=BLACK if c==4 else WHITE,
                 bg=MID_GRAY if c!=4 else ORANGE)
    note_cell(ws8, 4, c, hint)

ex8 = [
    ["FEV 2026", "Bruno", "2026-02-07", 14900, "confirmado", "Empresa XPTO", 1, "Pagamento à vista"],
    ["FEV 2026", "João",  "2026-02-08", 9800,  "na_rua",    "Startup Y",    2, "Boleto enviado"],
]
for r, row in enumerate(ex8, 5):
    for c, v in enumerate(row, 1):
        bg = LIGHT_GRAY
        if c == 5:
            bg = {"confirmado": "1A3320", "na_rua": "2A2010", "cancelado": "3A1A1A"}.get(v, LIGHT_GRAY)
        fmt = 'R$ #.##0,00' if c == 4 else None
        apply_cell(ws8, r, c, v, bg=bg, number_fmt=fmt)

dv8 = DataValidation(type="list", formula1='"confirmado,na_rua,cancelado"')
ws8.add_data_validation(dv8)
dv8.sqref = "E5:E500"

for col, w in zip("ABCDEFGH", [14, 12, 14, 14, 14, 24, 16, 28]):
    ws8.column_dimensions[col].width = w
ws8.row_dimensions[3].height = 22
ws8.row_dimensions[4].height = 14
freeze(ws8, "A5")


# ════════════════════════════════════════════════════════════════════════
# ABA 9 — GUIA RÁPIDO
# ════════════════════════════════════════════════════════════════════════
ws9 = wb.create_sheet("📖 Guia Rápido")
tab_color(ws9, "333333")

ws9.column_dimensions["A"].width = 22
ws9.column_dimensions["B"].width = 60

def guide_title(row, text):
    ws9.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    c = ws9.cell(row=row, column=1, value=text)
    c.fill = fill(BLACK)
    c.font = font(bold=True, color=ACCENT, size=12)
    c.alignment = left()
    ws9.row_dimensions[row].height = 22

def guide_row(row, label, value, highlight=False):
    c1 = ws9.cell(row=row, column=1, value=label)
    c1.fill = fill(DARK_GRAY)
    c1.font = font(bold=True, color=ACCENT2 if highlight else WHITE, size=10)
    c1.alignment = left()
    c2 = ws9.cell(row=row, column=2, value=value)
    c2.fill = fill(MID_GRAY)
    c2.font = font(color=WHITE, size=10)
    c2.alignment = left(wrap=True)
    ws9.row_dimensions[row].height = 20

guide_title(1,  "OXICORE REVENUE — GUIA RÁPIDO DE PREENCHIMENTO")
guide_title(3,  "FLUXO DIÁRIO")
guide_row(4,  "Manhã",         "✅ Adicionar linha em 📆 Dias Úteis  |  ✅ Lançar leads do dia anterior em 🔥 Leads")
guide_row(5,  "Durante o dia", "✅ SDR: criar linha em 📞 Reuniões ao marcar  |  ✅ Atualizar status após a reunião")
guide_row(6,  "Ao fechar",     "✅ Closer: criar linha em 💎 Vendas com status 'na_rua'")
guide_row(7,  "Pagamento",     "✅ Atualizar status de 'na_rua' para 'confirmado' em 💎 Vendas")
guide_row(8,  "Início do mês", "✅ Criar período em 📅 Períodos  |  ✅ Definir metas em 🎯 Metas Mensais  |  ✅ Lançar verba em 💰 Investimento")

guide_title(10, "STATUS DE REUNIÃO")
guide_row(11, "marcada",   "Agendada, ainda não aconteceu", True)
guide_row(12, "show",      "Aconteceu — cliente apareceu", True)
guide_row(13, "no_show",   "Não aconteceu — cliente sumiu", True)
guide_row(14, "remarcada", "Cliente pediu para reagendar", True)
guide_row(15, "cancelada", "Cancelada definitivamente", True)

guide_title(17, "STATUS DE VENDA")
guide_row(18, "confirmado", "Pagamento recebido — conta na conta")
guide_row(19, "na_rua",     "Proposta aceita, aguardando pagamento")
guide_row(20, "cancelado",  "Não concretizou / chargeback")

guide_title(22, "MÉTRICAS CALCULADAS AUTOMATICAMENTE (no banco)")
guide_row(23, "CPL",         "Investimento ÷ Leads")
guide_row(24, "CPC",         "Investimento ÷ Reuniões Marcadas")
guide_row(25, "CPR",         "Investimento ÷ Reuniões Acontecidas")
guide_row(26, "CPV",         "Investimento ÷ Vendas")
guide_row(27, "ROAS",        "Receita ÷ Investimento")
guide_row(28, "Ticket Médio","Receita ÷ Vendas")
guide_row(29, "Pacing",      "(Meta / Dias úteis mês) × Dias úteis hoje  →  compara com receita real")


# ════════════════════════════════════════════════════════════════════════
# Salvar
# ════════════════════════════════════════════════════════════════════════
out = "/home/oxicoreco/Documentos/GitHub/oxicorerevenue-main/docs/oxicore_revenue_database.xlsx"
wb.save(out)
print(f"Salvo em: {out}")
